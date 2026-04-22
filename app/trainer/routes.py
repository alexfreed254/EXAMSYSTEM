from flask import render_template, redirect, url_for, flash, request, send_file, g
from functools import wraps
from app.trainer import trainer
from app.models import (Trainer, Trainee, Course, Result, AcademicYear,
                         Program, Enrollment, User, Notification)
from app import db
from app.auth.routes import login_required
import pandas as pd
import os
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def trainer_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = g.get('current_user')
        if not user:
            return redirect(url_for('auth.login'))
        if user.role != 'trainer':
            flash('Trainer access required.', 'danger')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated


def get_trainer():
    return Trainer.query.filter_by(user_id=g.current_user.id).first()


@trainer.route('/dashboard')
@login_required
@trainer_required
def dashboard():
    t = get_trainer()
    if not t:
        flash('Trainer profile not found.', 'danger')
        return redirect(url_for('auth.login'))

    stats = {
        'total_results': Result.query.filter_by(trainer_id=t.id).count(),
        'published': Result.query.filter_by(trainer_id=t.id, is_published=True).count(),
        'pending': Result.query.filter_by(trainer_id=t.id, is_published=False).count(),
    }
    recent_results = Result.query.filter_by(trainer_id=t.id).order_by(
        Result.uploaded_at.desc()).limit(10).all()
    return render_template('trainer/dashboard.html', trainer=t, stats=stats,
                           recent_results=recent_results)


@trainer.route('/upload-results', methods=['GET', 'POST'])
@login_required
@trainer_required
def upload_results():
    t = get_trainer()
    courses = Course.query.filter_by(is_active=True).order_by(Course.name).all()
    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    current_year = AcademicYear.query.filter_by(is_current=True).first()

    if request.method == 'POST':
        course_id = request.form.get('course_id', type=int)
        year_id = request.form.get('year_id', type=int)
        file = request.files.get('excel_file')

        if not file or not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload a valid Excel file (.xlsx or .xls).', 'danger')
            return render_template('trainer/upload_results.html', trainer=t,
                                   courses=courses, years=years, current_year=current_year)

        course = Course.query.get_or_404(course_id)

        try:
            df = pd.read_excel(file)
            df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]

            required_cols = ['admission_number']
            missing = [c for c in required_cols if c not in df.columns]
            if missing:
                flash(f'Missing columns: {", ".join(missing)}. See template for format.', 'danger')
                return render_template('trainer/upload_results.html', trainer=t,
                                       courses=courses, years=years, current_year=current_year)

            success_count = 0
            error_rows = []

            for idx, row in df.iterrows():
                admission_no = str(row.get('admission_number', '')).strip()
                trainee = Trainee.query.filter_by(admission_number=admission_no).first()

                if not trainee:
                    error_rows.append(f'Row {idx+2}: Admission number "{admission_no}" not found.')
                    continue

                # Get marks (all out of 100)
                ca_theory = _safe_float(row.get('ca_theory'))
                ca_practical = _safe_float(row.get('ca_practical'))
                sa_theory = _safe_float(row.get('sa_theory'))
                sa_practical = _safe_float(row.get('sa_practical'))

                # Validate marks are 0-100
                for val, name in [(ca_theory, 'CA Theory'), (ca_practical, 'CA Practical'),
                                   (sa_theory, 'SA Theory'), (sa_practical, 'SA Practical')]:
                    if val is not None and (val < 0 or val > 100):
                        error_rows.append(f'Row {idx+2}: {name} must be 0-100.')
                        continue

                # Upsert result
                result = Result.query.filter_by(
                    trainee_id=trainee.id,
                    course_id=course_id,
                    academic_year_id=year_id
                ).first()

                if not result:
                    result = Result(
                        trainee_id=trainee.id,
                        course_id=course_id,
                        academic_year_id=year_id,
                        trainer_id=t.id,
                    )
                    db.session.add(result)

                result.ca_theory = ca_theory
                result.ca_practical = ca_practical
                result.sa_theory = sa_theory
                result.sa_practical = sa_practical
                result.trainer_id = t.id
                result.compute_results()
                success_count += 1

            db.session.commit()

            if error_rows:
                flash(f'Uploaded {success_count} results. Errors: ' + '; '.join(error_rows[:5]), 'warning')
            else:
                flash(f'Successfully uploaded {success_count} results for {course.name}.', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'Error processing file: {str(e)}', 'danger')

        return redirect(url_for('trainer.my_results'))

    return render_template('trainer/upload_results.html', trainer=t,
                           courses=courses, years=years, current_year=current_year)


def _safe_float(val):
    try:
        if val is None or str(val).strip() in ('', 'nan', 'NaN', 'N/A', '-'):
            return None
        return float(val)
    except (ValueError, TypeError):
        return None


@trainer.route('/download-template/<int:course_id>')
@login_required
@trainer_required
def download_template(course_id):
    course = Course.query.get_or_404(course_id)

    # Get enrolled trainees for this course's program
    trainees = (Trainee.query
                .join(Enrollment, Trainee.id == Enrollment.trainee_id)
                .filter(Enrollment.program_id == course.program_id)
                .join(User, Trainee.user_id == User.id)
                .order_by(User.full_name).all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Results'

    # Header styling
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells('A1:I1')
    ws['A1'] = f'TTTI Results Upload Template - {course.name} ({course.code})'
    ws['A1'].font = Font(bold=True, size=13, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:I2')
    ws['A2'] = f'Theory Weight: {course.theory_weight}% | Practical Weight: {course.practical_weight}% | All marks out of 100'
    ws['A2'].font = Font(italic=True, size=10, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = ['Admission Number', 'Full Name', 'CA Theory\n(out of 100)',
               'CA Practical\n(out of 100)', 'SA Theory\n(out of 100)',
               'SA Practical\n(out of 100)', 'Remarks']
    col_widths = [20, 30, 18, 18, 18, 18, 25]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[3].height = 35

    # Data rows
    alt_fill = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')
    for row_idx, trainee in enumerate(trainees, 4):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        data = [trainee.admission_number, trainee.user.full_name, '', '', '', '', '']
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            if fill.fill_type:
                cell.fill = fill

    # Freeze header
    ws.freeze_panes = 'A4'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'results_template_{course.code}_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@trainer.route('/results')
@login_required
@trainer_required
def my_results():
    t = get_trainer()
    course_id = request.args.get('course_id', type=int)
    year_id = request.args.get('year_id', type=int)

    query = Result.query.filter_by(trainer_id=t.id)
    if course_id:
        query = query.filter_by(course_id=course_id)
    if year_id:
        query = query.filter_by(academic_year_id=year_id)

    results_list = query.order_by(Result.uploaded_at.desc()).all()
    courses = Course.query.filter_by(is_active=True).order_by(Course.name).all()
    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()

    return render_template('trainer/results.html', trainer=t, results=results_list,
                           courses=courses, years=years,
                           selected_course=course_id, selected_year=year_id)


@trainer.route('/results/edit/<int:result_id>', methods=['GET', 'POST'])
@login_required
@trainer_required
def edit_result(result_id):
    t = get_trainer()
    result = Result.query.get_or_404(result_id)

    if result.trainer_id != t.id:
        flash('You can only edit your own results.', 'danger')
        return redirect(url_for('trainer.my_results'))

    if result.is_published:
        flash('Published results cannot be edited.', 'warning')
        return redirect(url_for('trainer.my_results'))

    if request.method == 'POST':
        result.ca_theory = _safe_float(request.form.get('ca_theory'))
        result.ca_practical = _safe_float(request.form.get('ca_practical'))
        result.sa_theory = _safe_float(request.form.get('sa_theory'))
        result.sa_practical = _safe_float(request.form.get('sa_practical'))
        result.compute_results()
        db.session.commit()
        flash('Result updated.', 'success')
        return redirect(url_for('trainer.my_results'))

    return render_template('trainer/edit_result.html', result=result, trainer=t)


@trainer.route('/trainees')
@login_required
@trainer_required
def trainees():
    t = get_trainer()
    # Show trainees in trainer's department programs
    if t.department_id:
        programs = Program.query.filter_by(department_id=t.department_id).all()
        program_ids = [p.id for p in programs]
        trainees_list = (Trainee.query
                         .join(Enrollment, Trainee.id == Enrollment.trainee_id)
                         .filter(Enrollment.program_id.in_(program_ids))
                         .join(User, Trainee.user_id == User.id)
                         .order_by(User.full_name).all())
    else:
        trainees_list = Trainee.query.join(User).order_by(User.full_name).all()

    return render_template('trainer/trainees.html', trainer=t, trainees=trainees_list)


@trainer.route('/profile')
@login_required
@trainer_required
def profile():
    t = get_trainer()
    return render_template('trainer/profile.html', trainer=t)
